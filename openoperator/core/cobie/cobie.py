
import pandas as pd
import rdflib
from rdflib import Namespace, Literal
import urllib.parse
import openpyxl
from openpyxl.styles import PatternFill

# Define common namespaces
COBIE = Namespace("http://checksem.u-bourgogne.fr/ontology/cobie24#")
RDF = Namespace("http://www.w3.org/1999/02/22-rdf-syntax-ns#")
A = RDF.type

class COBie:
    """
    This class handles everything related to the COBie graph.

    Its repsonsibilities are to:

    1. COBie spreadsheet validation
    2. Spreadsheet to RDF conversion
    """
    def __init__(self, file_path: str) -> None:
        self.file_path = file_path

    def create_uri(self, name: str) -> str:
        """
        Create a URI from string.
        """
        # name = re.sub(r'[^a-zA-Z0-9]', '', str(name).lower())
        # name = name.replace("'", "_")  # Replace ' with _
        name = urllib.parse.quote(name.lower())
        return name

    def validate_spreadsheet(self) -> (bool, dict, bytes):
        """
        Validate a COBie spreadsheet. Refer to COBie_validation.pdf in docs/ for more information.
        
        Returns: 

        - errors_found: A boolean indicating whether or not errors were found in the spreadsheet.
        - errors: A dictionary containing all the errors found in the spreadsheet.
        - updated_file: The file path of the updated spreadsheet with the errors highlighted in red.
        """
        errors = {
            "Expected sheet not found in spreadsheet.": [],
            "More than one record found in Facility sheet.": [],
            "Empty or N/A cells found in column A of sheet.": [],
            "Duplicate names found in column A of sheet.": [],
            "Space is not linked to a value in the first column of the Floor tab.": [],
            "Not every Type record has a category.": [],
            "Component is not linked to an existing Type.": [],
            "Component is not linked to an existing Space.": []
        }

        errors_found = False

        # Open COBie spreadsheet
        df = pd.read_excel(self.file_path, engine='openpyxl', sheet_name=None) 
        wb = openpyxl.load_workbook(self.file_path)

        expected_sheets = ['Facility', 'Floor', 'Space', 'Type', 'Component', 'Attribute', 'System']
        # Check to make sure the spreadsheet has the correct sheets     
        for sheet in expected_sheets:
            if sheet not in df.keys():
                errors["Expected sheet not found in spreadsheet."].append({
                    "sheet": sheet,
                })
                errors_found = True

                # highlight the sheet in red
                cell = wb[sheet].cell(row=1, column=1)
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

        # Make sure there is only one record in the Facility sheet
        if len(df['Facility']) > 1:
            errors["More than one record found in Facility sheet."].append({
                "sheet": "Facility",
                "row": 1,
                "column": 1
            })
            errors_found = True

            # highlight the sheet in red
            cell = wb['Facility'].cell(row=1, column=1)
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

        # No empty or N/A cells are present in column A of any sheet
        for sheet in expected_sheets:
            for idx, value in enumerate(df[sheet]['Name']):
                if pd.isnull(value):
                    errors["Empty or N/A cells found in column A of sheet."].append({
                        "sheet": sheet,
                        "row": idx + 2,
                        "column": 1
                    })
                    errors_found = True

                    # highlight the cell in red
                    cell = wb[sheet].cell(row=idx + 2, column=1)
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

        # Check Floor, Space, Type, Component sheets for duplicate names in column A
        for sheet in ['Floor', 'Space', 'Type', 'Component']:
            for idx, name in enumerate(df[sheet]['Name']):
                if name in df[sheet]['Name'][idx + 1:]:
                    errors["Duplicate names found in column A of sheet."].append({
                        "sheet": sheet,
                        "row": idx + 2,
                        "column": 1
                    })  
                    errors_found = True

                    cell = wb[sheet].cell(row=idx + 2, column=1)
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

        # Space Tab
        # Every value is linked to a value in the first column of the Floor tab
        for idx, space in enumerate(df['Space']['FloorName']):
            if space not in df['Floor']['Name'].values:
                errors["Space is not linked to a value in the first column of the Floor tab."].append({
                    "sheet": "Space",
                    "row": idx + 2,
                    "column": 5
                })  
                errors_found = True

                cell = wb['Space'].cell(row=idx + 2, column=5)
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

        # Type Tab
        # Every record has a category
        for idx, type in enumerate(df['Type']['Category']):
            if pd.isnull(type):
                errors["Not every Type record has a category."].append({
                    "sheet": "Type",
                    "row": idx + 2,
                    "column": 4
                })
                errors_found = True

                cell = wb['Type'].cell(row=idx + 2, column=4)
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
            
        # Component Tab
        # Every record is linked to a existing Type
        for idx, component in enumerate(df['Component']['TypeName']):
            if component not in df['Type']['Name'].values:
                errors["Component is not linked to an existing Type."].append({
                    "sheet": "Component",
                    "row": idx + 2,
                    "column": 4
                })
                errors_found = True

                cell = wb['Component'].cell(row=idx + 2, column=4)
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

        # Every record is linked a to existing Space
        for idx, space_name in enumerate(df['Component']['Space']):
            # Check if cell is valid
            if pd.isnull(space_name):
                errors["Component is not linked to an existing Space."].append({
                    "sheet": "Component",
                    "row": idx + 2,
                    "column": 5
                })
                errors_found = True

                cell = wb['Component'].cell(row=idx + 2, column=5)
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                continue
            # Split by "," to get all spaces and remove whitespace
            spaces = [space.strip() for space in space_name.split(",")]
            for space in spaces:
                if space not in df['Space']['Name'].values:
                    errors["Component is not linked to an existing Space."].append({
                        "sheet": "Component",
                        "row": idx + 2,
                        "column": 5
                    })
                    errors_found = True

                    cell = wb['Component'].cell(row=idx + 2, column=5)
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

        # Remove all the empty lists from the errors dict
        errors = {key: value for key, value in errors.items() if value}

        # Save the workbook
        wb.save("test.xlsx")

        return errors_found, errors, "test.xlsx"

    def convert_to_graph(self, namespace: str) -> str:
        """
        Converts a valid COBie spreadsheet to RDF and uploads it to knowledge graph.

        portfolio_namespace: The namespace to represent a group of buildings and is used to create the URI of the facility. Ex. https://departmentOfEnergy.com/ could be the namespace for all the buildings owned by the Department of Energy. 
        """
        # Make sure the portfolio namespace is a valid URI
        assert urllib.parse.urlparse(namespace).scheme != ""
        namespace = Namespace(namespace)

        # Open COBie spreadsheet
        df = pd.read_excel(self.file_path, engine='openpyxl', sheet_name=None)

        errors_found, errors, updated_file_path = self.validate_spreadsheet()

        for key, value in errors.items():
            if len(value) > 0:
                print(key)
                print(value)

        if errors_found:
            print("Errors found in the spreadsheet:")
            return errors
        else:
            print("No errors found in the spreadsheet.")

            # Create an rdflib Graph to store the RDF data
            g = rdflib.Graph()
            g.bind("RDF", RDF)
            g.bind("COBIE", COBIE)
            g.bind("Namespace", namespace)

            facility_uri = namespace # All the nodes in the graph will extend this uri

            for sheet in ['Floor', 'Space', 'Type', 'Component', 'System']:
                print(f"Processing {sheet} sheet...")
                predicates = df[sheet].keys()
                predicates = [predicate[0].lower() + predicate[1:] for predicate in predicates] # Make first letter lowercase

                # Iterate over all rows in the sheet, skipping the first row
                for _, row in df[sheet].iterrows():
                    # The name field is used as the subject
                    subject = row['Name']
                    subject_uri = facility_uri["/" + sheet.lower() + "/" + self.create_uri(subject)]

                    # Get the values of the row
                    objects = row.values

                    # Create the node
                    g.add((subject_uri, A, COBIE[sheet]))

                    # Add objects
                    i = 0
                    for obj in objects:
                        predicate = predicates[i]

                        # Check if it should be a relationship or a literal
                        if (sheet == "Component" and predicate == "typeName") or (sheet == "Space" and predicate == "floorName") or (sheet == "System" and predicate == "componentNames"):
                            # The target sheet is the one where the relationship is pointing to 
                            target_sheet = None
                            if sheet == "Component":
                                target_sheet = "Type"
                            elif sheet == "Space":
                                target_sheet = "Floor"
                            elif sheet == "System":
                                target_sheet = "Component"

                            g.add((subject_uri, COBIE[predicate], facility_uri["/" + target_sheet.lower() + "/" + self.create_uri(obj)]))
                        elif sheet == "Component" and predicate == "space":
                            # Split by "," to get all spaces and remove whitespace
                            spaces = [space.strip() for space in obj.split(",")]
                            for space in spaces:
                                g.add((subject_uri, COBIE[predicate], facility_uri["/" + "space" + "/" + self.create_uri(space)]))
                        else:
                            g.add((subject_uri, COBIE[predicate], Literal(str(obj).replace('"', '\\"'))))
                        i += 1      

            # Create the attributes
            print("Processing Attribute sheet...")
            for _, row in df['Attribute'].iterrows():
                target_sheet = row['SheetName']
                target_row_name = row['RowName']
                target_uri = facility_uri["/" + target_sheet.lower() + "/" + self.create_uri(target_row_name)]

                attribute_uri = target_uri + "/attribute/" + self.create_uri(row['Name'])

                g.add((attribute_uri, A, COBIE['Attribute']))
                g.add((attribute_uri, COBIE['name'], Literal(row['Name'])))
                g.add((attribute_uri, COBIE['value'], Literal(row['Value'])))
                g.add((attribute_uri, COBIE['unit'], Literal(row['Unit'])))
                g.add((attribute_uri, COBIE['attributeTo'], target_uri))

            # Serialize the graph to a file
            graph_string = g.serialize(format='turtle', encoding='utf-8').decode()

            return graph_string
