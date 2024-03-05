from typing import Dict, Tuple
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO

class COBieSpreadsheet:
  def __init__(self, file_content: bytes):
    self.file_content = file_content

  def validate(self) -> Tuple[bool, Dict, bytes]:
    """
    Validate a COBie spreadsheet. Refer to COBie_validation.pdf in docs/ for more information.

    Returns: 
    - errors_found: A boolean indicating whether or not errors were found in the spreadsheet.
    - errors: A dictionary containing all the errors found in the spreadsheet.
    - updated_file: The file path of the updated spreadsheet with the errors highlighted in red.
    """
    errors = {
        "Expected sheet not found in spreadsheet.": [],
        "More than one record found in Facility sheet.": [],
        "Empty or N/A cells found in column A of sheet.": [],
        "Duplicate names found in column A of sheet.": [],
        "Space is not linked to a value in the first column of the Floor tab.": [],
        "Not every Type record has a category.": [],
        "Component is not linked to an existing Type.": [],
        "Component is not linked to an existing Space.": []
    }

    errors_found = False

    # Open COBie spreadsheet
    df = pd.read_excel(BytesIO(self.file_content), engine='openpyxl', sheet_name=None) 
    wb = openpyxl.load_workbook(BytesIO(self.file_content))

    expected_sheets = ['Facility', 'Floor', 'Space', 'Type', 'Component', 'Attribute', 'System']
    # Check to make sure the spreadsheet has the correct sheets     
    for sheet in expected_sheets:
      if sheet not in df.keys():
        errors["Expected sheet not found in spreadsheet."].append({
            "sheet": sheet,
        })
        errors_found = True
    if errors_found: return errors_found, errors, self.file_content

    # Make sure there is only one record in the Facility sheet
    if len(df['Facility']) > 1:
      errors["More than one record found in Facility sheet."].append({
        "sheet": "Facility",
        "row": 1,
        "column": 1
      })
      errors_found = True

      # highlight the sheet in red
      cell = wb['Facility'].cell(row=1, column=1)
      cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

    # No empty or N/A cells are present in column A of any sheet
    for sheet in expected_sheets:
      for idx, value in enumerate(df[sheet]['Name']):
        if pd.isnull(value) or value == None or value == "N/A" or pd.isna(value):
          errors["Empty or N/A cells found in column A of sheet."].append({
            "sheet": sheet,
            "row": idx + 2,
            "column": 1
          })
          errors_found = True

          # highlight the cell in red
          cell = wb[sheet].cell(row=idx + 2, column=1)
          cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

    # Check Floor, Space, Type, Component sheets for duplicate names in column A
    for sheet in ['Floor', 'Space', 'Type', 'Component']:
      seen_names = set()
      for idx, name in enumerate(df[sheet]['Name']):
        if name in seen_names:
          errors["Duplicate names found in column A of sheet."].append({
            "sheet": sheet,
            "row": idx + 2,
            "column": 1
          })  
          errors_found = True

          cell = wb[sheet].cell(row=idx + 2, column=1)
          cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
        else:
          seen_names.add(name)

    # Space Tab
    # Every value is linked to a value in the first column of the Floor tab
    for idx, space in enumerate(df['Space']['FloorName']):
      if space not in df['Floor']['Name'].values:
        errors["Space is not linked to a value in the first column of the Floor tab."].append({
          "sheet": "Space",
          "row": idx + 2,
          "column": 5
        })  
        errors_found = True

        cell = wb['Space'].cell(row=idx + 2, column=5)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

    # Type Tab
    # Every record has a category
    for idx, type in enumerate(df['Type']['Category']):
      if pd.isnull(type):
        errors["Not every Type record has a category."].append({
          "sheet": "Type",
          "row": idx + 2,
          "column": 4
        })
        errors_found = True

        cell = wb['Type'].cell(row=idx + 2, column=4)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
        
    # Component Tab
    # Every record is linked to a existing Type
    for idx, component in enumerate(df['Component']['TypeName']):
      if component not in df['Type']['Name'].values:
        errors["Component is not linked to an existing Type."].append({
          "sheet": "Component",
          "row": idx + 2,
          "column": 4
        })
        errors_found = True

        cell = wb['Component'].cell(row=idx + 2, column=4)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

    # Every record is linked a to existing Space
    for idx, space_name in enumerate(df['Component']['Space']):
      # Check if cell is valid
      if pd.isnull(space_name):
        errors["Component is not linked to an existing Space."].append({
          "sheet": "Component",
          "row": idx + 2,
          "column": 5
        })
        errors_found = True

        cell = wb['Component'].cell(row=idx + 2, column=5)
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
        continue
      # Split by "," to get all spaces and remove whitespace
      spaces = [space.strip() for space in space_name.split(",")]
      for space in spaces:
        if space not in df['Space']['Name'].values:
          errors["Component is not linked to an existing Space."].append({
            "sheet": "Component",
            "row": idx + 2,
            "column": 5
          })
          errors_found = True

          cell = wb['Component'].cell(row=idx + 2, column=5)
          cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")

    # Remove all the empty lists from the errors dict
    errors = {key: value for key, value in errors.items() if value}

    # Save the workbook
    content = BytesIO()
    wb.save(content)
    content.seek(0)

    return errors_found, errors, content.getvalue()