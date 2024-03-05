import unittest
from openoperator.domain.model import COBieSpreadsheet
from openpyxl import Workbook
from io import BytesIO

class TestCOBie(unittest.TestCase):
  # def setUp(self) -> None:
  #   self.facility_mock = Mock()
  #   self.facility_mock.uri = "https://openoperator.com/facility"
  #   self.embeddings_mock = Mock()
  #   self.cobie = COBieSpreadsheet(facility=self.facility_mock, embeddings=self.embeddings_mock)
  
  def create_cobie_spreadsheet(self, modifications=None) -> COBieSpreadsheet:
    """
    Create a basic COBie spreadsheet and apply modifications if provided.
    Modifications should be a dict with sheet names as keys and lists of rows as values.
    """
    wb = Workbook()
    sheets_data = {
        "Facility": [
            ["Name"],
            ["Test Facility"]
        ],
        "Floor": [
            ["Name"],
            ["Test Floor"],
            ["Test Floor 2"],
        ],
        "Space": [
            ["Name", "CreatedBy", "CreatedOn", "Category", "FloorName"],
            ["Test Space", "Test User", "2022-01-01", "Space", "Test Floor"],
            ["Test Space 2", "Test User", "2022-01-01", "Space", "Test Floor 2"]
        ],
        "Type": [
            ["Name", "CreatedBy", "CreatedOn", "Category"],
            ["Test Door", "Test User", "2022-01-01", "23-30 10: Doors"],
        ],
        "Component": [
            ["Name", "CreatedBy", "CreatedOn", "TypeName", "Space"],
            ["Test Component", "Test User", "2022-01-01", "Test Door", "Test Space"],
            ["Test Component 2", "Test User", "2022-01-01", "Test Door", "Test Space 2"]
        ],
        "System": [
            ["Name", "CreatedBy", "CreatedOn", "Category", "ComponentNames"],
            ["Test System", "Test User", "2022-01-01", "System", "Test Component"],
        ],
        "Attribute": [
            ["Name", "CreatedBy", "CreatedOn", "Category", "SheetName", "RowName", "Value", "Unit"],
            ["Test Attribute", "Test User", "2022-01-01", "Attribute", "Component", "Test Component", "Test Value", "Test Unit"],
            ["Test Attribute 2", "Test User", "2022-01-01", "Attribute", "Component", "Test Component 2", "Test Value", "Test Unit"],
        ],
    }

    for sheet, rows in sheets_data.items():
      if sheet not in wb.sheetnames:
        wb.create_sheet(sheet)
      ws = wb[sheet]
      for row in rows:
        ws.append(row)

    # Apply modifications
    if modifications:
      for sheet, rows in modifications.items():
        ws = wb[sheet]
        for row in rows:
          ws.append(row)

    file_content = BytesIO()
    wb.save(file_content)
    file_content.seek(0)
    spreadsheet = COBieSpreadsheet(file_content=file_content.getvalue())
    return spreadsheet

  def test_validate_spreadsheet_success(self):
    spreadsheet = self.create_cobie_spreadsheet()
    errors_found, errors, _ = spreadsheet.validate()
    assert errors_found == False
    assert errors == {}

  def test_validate_spreadsheet_multiple_facility_records(self):
    """Test for multiple records in Facility sheet."""
    spreadsheet = self.create_cobie_spreadsheet(modifications={
      "Facility": [["Name2", "Value2"]]
    })
    errors_found, errors, _ = spreadsheet.validate()
    assert errors_found == True
    assert "More than one record found in Facility sheet." in errors

  def test_validate_spreadsheet_empty_cells_in_column_A(self):
    """Test for empty or N/A cells in column A of any sheet."""
    modifications = {
        "Floor": [[None], # Adding an empty row to Floor
                  ['Floor 5']]  
    }
    spreadsheet = self.create_cobie_spreadsheet(modifications=modifications)
    errors_found, errors, _ = spreadsheet.validate()
    assert errors_found == True
    assert "Empty or N/A cells found in column A of sheet." in errors

  def test_validate_spreadsheet_duplicate_names_in_column_A(self):
    """Test for duplicate names in column A of specific sheets."""
    modifications = {
        "Component": [["Test Component 2", "Test User", "2022-01-01", "Test Door", "Test Space 2"]]  # Duplicate Component1
    }
    spreadsheet = self.create_cobie_spreadsheet(modifications=modifications)
    errors_found, errors, _ = spreadsheet.validate()
    assert errors_found == True
    assert "Duplicate names found in column A of sheet." in errors

  def test_validate_spreadsheet_unlinked_space_records(self):
    """Test for unlinked Space records."""
    modifications = {
        "Space": [["Test Space 2", "Test User", "2022-01-01", "Space", "Floor 999"]]  # Unlinked Space
    }
    spreadsheet = self.create_cobie_spreadsheet(modifications=modifications)
    errors_found, errors, _ = spreadsheet.validate()
    assert errors_found == True
    assert "Space is not linked to a value in the first column of the Floor tab." in errors

  def test_validate_spreadsheet_type_no_category(self):
    modifications = {
      "Type": [["Test No Category", "Test User", "2022-01-01", ""]]  # Type with no category
    }
    spreadsheet = self.create_cobie_spreadsheet(modifications=modifications)
    errors_found, errors, _ = spreadsheet.validate()
    assert errors_found == True
    assert "Not every Type record has a category." in errors

  def test_validate_spreadsheet_component_no_type(self):
    modifications = {
      "Component": [["Test Component No Type", "Test User", "2022-01-01", "", "Test Space 2"]]  # Component with no type
    }
    spreadsheet = self.create_cobie_spreadsheet(modifications=modifications)
    errors_found, errors, _ = spreadsheet.validate()
    assert errors_found == True
    assert "Component is not linked to an existing Type." in errors

  def test_validate_spreadsheet_component_no_space(self):
    modifications = {
      "Component": [["Test Component No Space", "Test User", "2022-01-01", "Test Door", ""]]  # Component with no space
    }
    spreadsheet = self.create_cobie_spreadsheet(modifications=modifications)
    errors_found, errors, _ = spreadsheet.validate()
    assert errors_found == True
    assert "Component is not linked to an existing Space." in errors   

if __name__ == "__main__":
  unittest.main()