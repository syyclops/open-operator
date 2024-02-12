from unittest.mock import Mock, patch
import unittest
from openoperator.core.cobie.cobie import COBie 
from openpyxl import Workbook
from io import BytesIO
from rdflib import Graph

class TestCOBie(unittest.TestCase):
    def setUp(self) -> None:
        self.facility_mock = Mock()
        self.facility_mock.uri = "https://openoperator.com/facility"
        self.embeddings_mock = Mock()
        self.cobie = COBie(facility=self.facility_mock, embeddings=self.embeddings_mock)
    
    def create_cobie_spreadsheet(self, modifications=None):
        """
        Create a basic COBie spreadsheet and apply modifications if provided.
        Modifications should be a dict with sheet names as keys and lists of rows as values.
        """
        wb = Workbook()
        sheets_data = {
            "Facility": [
                ["Name"],
                ["Test Facility"]
            ],
            "Floor": [
                ["Name"],
                ["Test Floor"],
                ["Test Floor 2"],
            ],
            "Space": [
                ["Name", "CreatedBy", "CreatedOn", "Category", "FloorName"],
                ["Test Space", "Test User", "2022-01-01", "Space", "Test Floor"],
                ["Test Space 2", "Test User", "2022-01-01", "Space", "Test Floor 2"]
            ],
            "Type": [
                ["Name", "CreatedBy", "CreatedOn", "Category"],
                ["Test Door", "Test User", "2022-01-01", "23-30 10: Doors"],
            ],
            "Component": [
                ["Name", "CreatedBy", "CreatedOn", "TypeName", "Space"],
                ["Test Component", "Test User", "2022-01-01", "Test Door", "Test Space"],
                ["Test Component 2", "Test User", "2022-01-01", "Test Door", "Test Space 2"]
            ],
            "System": [
                ["Name", "CreatedBy", "CreatedOn", "Category", "ComponentNames"],
                ["Test System", "Test User", "2022-01-01", "System", "Test Component"],
            ],
            "Attribute": [
                ["Name", "CreatedBy", "CreatedOn", "Category", "SheetName", "RowName", "Value", "Unit"],
                ["Test Attribute", "Test User", "2022-01-01", "Attribute", "Component", "Test Component", "Test Value", "Test Unit"],
                ["Test Attribute 2", "Test User", "2022-01-01", "Attribute", "Component", "Test Component 2", "Test Value", "Test Unit"],
            ],
        }

        for sheet, rows in sheets_data.items():
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
            ws = wb[sheet]
            for row in rows:
                ws.append(row)

        # Apply modifications
        if modifications:
            for sheet, rows in modifications.items():
                ws = wb[sheet]
                for row in rows:
                    ws.append(row)

        file_content = BytesIO()
        wb.save(file_content)
        file_content.seek(0)
        return file_content.getvalue()

    def test_validate_spreadsheet_success(self):
        file_content = self.create_cobie_spreadsheet()
        errors_found, errors, _ = self.cobie.validate_spreadsheet(file_content)
        assert errors_found == False
        assert errors == {}

    def test_validate_spreadsheet_multiple_facility_records(self):
        """Test for multiple records in Facility sheet."""
        file_content = self.create_cobie_spreadsheet(modifications={
            "Facility": [["Name2", "Value2"]]
        })
        errors_found, errors, _ = self.cobie.validate_spreadsheet(file_content=file_content)
        assert errors_found == True
        assert "More than one record found in Facility sheet." in errors

    def test_validate_spreadsheet_empty_cells_in_column_A(self):
        """Test for empty or N/A cells in column A of any sheet."""
        modifications = {
            "Floor": [[None], # Adding an empty row to Floor
                      ['Floor 5']]  
        }
        file_content = self.create_cobie_spreadsheet(modifications=modifications)
        errors_found, errors, _ = self.cobie.validate_spreadsheet(file_content=file_content)
        assert errors_found == True
        assert "Empty or N/A cells found in column A of sheet." in errors

    def test_validate_spreadsheet_duplicate_names_in_column_A(self):
        """Test for duplicate names in column A of specific sheets."""
        modifications = {
            "Component": [["Test Component 2", "Test User", "2022-01-01", "Test Door", "Test Space 2"]]  # Duplicate Component1
        }
        file_content = self.create_cobie_spreadsheet(modifications=modifications)
        errors_found, errors, _ = self.cobie.validate_spreadsheet(file_content=file_content)
        assert errors_found == True
        assert "Duplicate names found in column A of sheet." in errors

    def test_validate_spreadsheet_unlinked_space_records(self):
        """Test for unlinked Space records."""
        modifications = {
            "Space": [["Test Space 2", "Test User", "2022-01-01", "Space", "Floor 999"]]  # Unlinked Space
        }
        file_content = self.create_cobie_spreadsheet(modifications=modifications)
        errors_found, errors, _ = self.cobie.validate_spreadsheet(file_content=file_content)
        assert errors_found == True
        assert "Space is not linked to a value in the first column of the Floor tab." in errors

    def test_validate_spreadsheet_type_no_category(self):
        modifications = {
            "Type": [["Test No Category", "Test User", "2022-01-01", ""]]  # Type with no category
        }
        file_content = self.create_cobie_spreadsheet(modifications=modifications)
        errors_found, errors, _ = self.cobie.validate_spreadsheet(file_content=file_content)
        assert errors_found == True
        assert "Not every Type record has a category." in errors

    def test_validate_spreadsheet_component_no_type(self):
        modifications = {
            "Component": [["Test Component No Type", "Test User", "2022-01-01", "", "Test Space 2"]]  # Component with no type
        }
        file_content = self.create_cobie_spreadsheet(modifications=modifications)
        errors_found, errors, _ = self.cobie.validate_spreadsheet(file_content=file_content)
        assert errors_found == True
        assert "Component is not linked to an existing Type." in errors

    def test_validate_spreadsheet_component_no_space(self):
        modifications = {
            "Component": [["Test Component No Space", "Test User", "2022-01-01", "Test Door", ""]]  # Component with no space
        }
        file_content = self.create_cobie_spreadsheet(modifications=modifications)
        errors_found, errors, _ = self.cobie.validate_spreadsheet(file_content=file_content)
        assert errors_found == True
        assert "Component is not linked to an existing Space." in errors

    def test_convert_to_graph_basic(self):
        file_content = self.create_cobie_spreadsheet()
        graph_string = self.cobie.convert_to_graph(file_content)

        # Load into a graph and check for expected nodes and relationships
        g = Graph()
        g.parse(data=graph_string, format="turtle")

        assert len(g) == 49
        assert graph_string is not None
        assert "https://openoperator.com/facility" in graph_string
        assert "Test Space" in graph_string
        assert "Test Space 2" in graph_string
        assert "Test Floor" in graph_string
        assert "Test Floor 2" in graph_string
        assert "Test Door" in graph_string
        assert "Test Component" in graph_string
        assert "Test Component 2" in graph_string
        
   

if __name__ == "__main__":
    unittest.main()